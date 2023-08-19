import inspect
import logging
from datetime import date
from openpyxl import Workbook, load_workbook


class Utils:

    def custom_logger(logLevel=logging.DEBUG):
        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)
        fh = logging.FileHandler("..\TestFramework\logs\\"+str(date.today())+".log")
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s : %(message)s', datefmt='%m/%d/%Y %I:%M:%S %p')
        fh.setFormatter(formatter)
        logger.addHandler(fh)
        return logger

    def read_from_excel(file_name, sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct + 1):
            row = []
            for j in range(1, col_ct + 1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist
